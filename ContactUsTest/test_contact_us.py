import requests
import pytest
import datetime
import json
from pydantic import ValidationError, BaseModel
from typing import List
import pyotp
import threading
import time 
import jwt
import openpyxl
import io
import configparser

config = configparser.ConfigParser()
config.read("config.ini")

AUTH_URL = "https://beta.tradefinder.in/api_be/auth_internal/"
BASE_URL = "https://beta.tradefinder.in/api_be/contact_us"

#-------------------------- Pydantic Models ----------------------------#

# {'payload': {'data': [['1744796400', 'hr@gmail.com', 'harshada', '8433544979', 'this is testing contact message'], ['1727689882', 'nayan@gmail.com', 'nayan', '7894561230', 'This is a testing message']]}
class Payload(BaseModel):
    data: List[List[str]]

# response status
class Response(BaseModel):
    status: str

# ------------------------- Token Management ------------------------- #

# SECRET_KEY = pyotp.random_base32()
SECRET_KEY = config['ACCESS TOKEN']['SECRET_KEY']

lock = threading.Lock()

token_data = {
    'jwttoken': None,
    'accesstoken': None
}

def is_token_expired(token):
    try:
        decoded = jwt.decode(token, options={"verify_signature": False})
        exp = decoded.get("exp")
        return time.time() > exp if exp else True
    except Exception as e:
        print(f"\nError decoding token: {e}\n")
        return True

# JWT Token Func
def get_jwttoken():
    try:
        response = requests.get(f'{AUTH_URL}/bot_auth_internal',
                                headers={'Authorization':config['JWT TOKEN']['AUTHORIZATION'], 'User-Agent': config['JWT TOKEN']['USER_AGENT']}
                                )
        #return "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpZCI6Ijg1MGVlNGU0MWMzZjExZjBhMzEwMjczYTJmOGFhMGFhIiwiZW1haWwiOiJtYXl1cnNhbG9raGU5MjAxQGdtYWlsLmNvbSIsInN0YXJ0IjoxNzQ0OTcxOTU4LjkzNDcwNSwiZXhwIjoxNzQ0OTgyNzU4LjkzNDcwNSwicGxhbiI6IkRJQU1PTkQiLCJ1c2VyX3R5cGUiOiJjbGllbnQiLCJhY2Nlc3MiOiJ7XCJtYXJrZXRfcHVsc2VcIjogMSwgXCJpbnNpZGVyX3N0cmF0ZWd5XCI6IDAsIFwic2VjdG9yX3Njb3BlXCI6IDAsIFwic3dpbmdfc3BlY3RydW1cIjogMCwgXCJvcHRpb25fY2xvY2tcIjogMCwgXCJvcHRpb25fYXBleFwiOiAwLCBcInBhcnRuZXJfcHJvZ3JhbVwiOiAwfSIsImFjY291bnRfZXhwIjoiMTc3NjE5MTQwMCIsImJyb2tlciI6IiJ9.eWhTmvlOfPSnbEWmUj2tsvnCr6zxZXnoz1Rg7IK679o"  
        print(f'\nJwttoken:{response.text}\n')
        return response.text
    except Exception as e:
        print(f'Auth Internal JWTTOKEN Error:{e}')

# ACCESS TOKEN Func
def get_accesstoken():
    access_token = pyotp.TOTP(SECRET_KEY, interval=30).now()
    print(f'\nAccess Token:{access_token}\n')
    return access_token

def refresh_tokens():
    while True:
        with lock:
            if token_data['jwttoken'] is None or is_token_expired(token_data['jwttoken']):
                print("\nRefreshing JWT Token\n")
                token_data['jwttoken'] = get_jwttoken()
            token_data['accesstoken'] = get_accesstoken()
        time.sleep(30)

token_refresher = threading.Thread(target=refresh_tokens, daemon=True)
token_refresher.start()

@pytest.fixture(scope="module")
def auth_headers():
    with lock:
        return { 
            'Jwttoken': token_data['jwttoken'],
            'Accesstoken': token_data['accesstoken']
        }

# ------------------------- Health Check --------------------------- #

def test_health():
    """
    Test /health endpoint to ensure service is running.
    """
    response = requests.get(f'{BASE_URL}/health')
    response_json = response.json()

    print(f"Health Status: {response_json}\n")

    assert response.status_code == 200
    assert response_json.get('status') == "OK"

# ------------------------- Admin Contact Read ------------------------- #

def test_contact_admin_read(auth_headers):
    """
    Test /log_read endpoint and validate response schema.

    Response Example:
    {
        "payload": {
            "data": [
                [
                    "1744797014",
                    "hr@gmail.com",
                    "harshada",
                    "8433544979",
                    "this is testing contact message"
                ],
                [
                    "1744796400",
                    "hr@gmail.com",
                    "harshada",
                    "8433544979",
                    "this is testing contact message"
                ],
                [
                    "1727689882",
                    "nayan@gmail.com",
                    "nayan",
                    "7894561230",
                    "This is a testing message"
                ]
            ]
        },
        "status": "SUCCESS"
    }
    """
    print("\nRunning Contact Admin Read...\n")

    response = requests.get(f'{BASE_URL}/admin_contact', headers=auth_headers)
    response_json = response.json()

    print(f"CONTACT ADMIN READ: {response_json}\n")

    assert response.status_code == 200
    assert isinstance(response_json['payload'], dict)
    assert isinstance(response_json['payload']['data'], list)

    # Response Validation
    try:
        response_model = Response(**response_json)
        assert response_model.status == 'SUCCESS', "API response status is not SUCCESS"
    except ValidationError as e:
        pytest.fail(f"Response schema validation error: {e}")

    # Validate Payload structure
    try:
        payload_model = Payload(**response_json['payload'])
    except ValidationError as e:
        pytest.fail(f"Payload schema validation failed: {e}")

# ------------------------- Create & Delete Contact Log ------------------------- #

def get_latest_ts(auth_headers):
    """
    Get latest timestamp from /admin_contact.
    """
    print("\nFetching latest timestamp...\n")

    response = requests.get(f'{BASE_URL}/admin_contact', headers=auth_headers)
    assert response.status_code == 200

    response_json = response.json()
    data_list = response_json.get("payload", {}).get("data", [])
    assert data_list, "No log entries found"

    ts = data_list[0][0] # Assuming first item is the latest
    print(f"Latest Timestamp: {ts}\n")
    return ts

@pytest.fixture(scope="module")
def created_log_ts(auth_headers):
    """
    Create a new contact log and return its timestamp.
    """
    input_data = {
        "name": "testuser",
        "email": "testuser@gmail.com",
        "contact": "1234567890",
        "message": "this is testing contact message"
    }

    print(f"\nCreating new contact log with: {input_data}\n")

    response = requests.post(f'{BASE_URL}/contact_us_crud',
                                  json=input_data,
                                  headers=auth_headers)

    response_json = response.json()

    print(f"Create Response: {response_json}\n")

    assert response.status_code == 200

    try:
        response_model = Response(**response_json)
        assert response_model.status == 'SUCCESS'
    except ValidationError as e:
        pytest.fail(f"Create log response validation error: {e}")

    return get_latest_ts(auth_headers)

def test_contact_create(created_log_ts):
    """
    Ensure contact log creation returns a valid timestamp.
    """
    assert created_log_ts is not None, "Timestamp is None – event might not have been created properly"
    assert isinstance(created_log_ts, str)
    print(f"Created log with timestamp: {created_log_ts}\n")

def test_contact_delete(created_log_ts, auth_headers):
    """
    Delete the created log using timestamp.
    """
    print(f"\nDeleting contact log with timestamp: {created_log_ts}\n")

    response = requests.delete(f'{BASE_URL}/admin_contact',
                                      json={"timestamp": created_log_ts},
                                      headers=auth_headers)

    response_json = response.json()

    print(f"Delete Response: {response_json}\n")

    assert response.status_code == 200

    try:
        response_model = Response(**response_json)
        assert response_model.status == 'SUCCESS', "API response status is not SUCCESS"
    except ValidationError as e:
        pytest.fail(f"Delete response validation error: {e}")

# ------------------------- Download Excel ------------------------- #

def test_download_excel(auth_headers):
    """
    Test for downloading and validating Excel file
    """
    print("\nTesting Excel Download...\n")

    url = f'{BASE_URL}/download_excel'
    response = requests.get(url, headers=auth_headers)

    print(f"Download Status: {response.status_code}")
    print(f"Content-Type: {response.headers['Content-Type']}\n")

    assert response.status_code == 200
    assert response.headers['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    try:
        workbook = openpyxl.load_workbook(io.BytesIO(response.content))
        sheet_names = workbook.sheetnames

        print(f"Excel file is valid. Sheets: {sheet_names}\n")

        sheet = workbook.active
        assert sheet.max_row > 1, "Not enough rows in Excel"
        assert sheet.max_column > 1, "Not enough columns in Excel"
        assert sheet.cell(row=1, column=1).value is not None, "Top-left cell is empty"

    except Exception as e:
        pytest.fail(f"Failed to open/read Excel file: {e}")

 
    ## Download Excel file
    # filename = 'downloaded_file.xlsx'
    # if response.status_code == 200:
    #     with open(filename, 'wb') as f:
    #         f.write(response.content)
    #     print(f"Excel file downloaded and saved as '{filename}'")
    # else:
    #     print(f"Failed to download file. Status code: {response.status_code}")